from flask import request,send_file,Response
import requests
import pandas as pd
import csv
import time
import openpyxl
import os
from openpyxl.styles import Font

from pathlib import Path
from urllib.parse import quote

ALLOWED_EXTENSIONS = set(['csv'])
###############
# main function
###############

# 將輸入的CSV檔案轉換成EXCEL輸出
# 此function 需接收兩個檔案 
# 1.file 輸入檔案 
# 2.consignment 托寄物名稱
# 最先的判斷為 接收到的 file 和 consignment 是否為空 ，若為空 則回傳錯誤訊息 {Input cannot be empty!}
# 之後判斷 file 的檔案是否為指定副檔名，若否則回傳錯誤訊息 {Input type error!}
# 若判斷皆為是才去執行檔案轉換的動作
# 檔案轉換執行步驟
# part 1 
# 儲存輸入檔案至指定資料夾中static/uploads/，並取名為input.csv
# part 2
# 開啟儲存檔案做轉換為Json格式
# part 3
# 將Json格式資料轉換為xlsx檔案
# part 4
# 回傳xlsx檔案
def csvToExcel():
    # get upload file and consignment
    file = request.form['fileUrl']
    consignment = request.form['consignment']

    csvUrlRequest = getCsvfromUrl(file)
    filename = os.path.basename(file)
    today = time.strftime("%Y%m%d",time.localtime(time.time()))
    filename = "HK_CROS_CIC_ORDERS_" + today + ".csv"
    
    
    if file and consignment:
        if allowedFileType(filename):
            if csvUrlRequest == 200:
                # part 1 
                # save input
                # file.save('static/uploads/'+ "input.csv")

                

                # part 2
                # upload file processing 
                with open('static/uploads/' + "input.csv", 'r', encoding = "utf-8-sig") as csvFile:
                    reader = csv.reader(csvFile) 
                    readerInJson = readerIntoJson(reader,consignment)
                # file.close()

                # error code
                if readerInJson == 403:
                    # input 資料有地址為空
                    return Response('{"msg":"Input address  empty!"}', status=403, mimetype='application/json')
                elif readerInJson == 404:
                    # input 格式錯誤 以每一筆資料長度是否為32判斷
                    return Response('{"msg":"Input format error!"}', status=404, mimetype='application/json')
                elif readerInJson == 501:
                    # 外部分割地址API 發生錯誤
                    return Response('{"msg":"Address Split API server error!"}', status=501, mimetype='application/json')
                # part 3
                # json into xlsx
                xlsxFile = jsonIntoExcel(readerInJson,filename)
                
                # part 4
                return xlsxFile
            else:
                return Response('{"msg":"File url request error!"}', status=csvUrlRequest, mimetype='application/json')
        else:
            return Response('{"msg":"Input type error!"}', status=401, mimetype='application/json')
    else:
        return Response('{"msg":"Input cannot be empty!"}', status=402, mimetype='application/json')


#################
# other function
#################


# request file from url
# 將input給的url去下載該檔案並儲存至 static/uploads/ 資料夾中
# part 1
# 發送請求將回傳值存入response中
# part 2
# error 判斷 ， 若發送請求status不為200，就當作發送請求失敗，回傳請求後的error status code
# part 3
# 發送成功將input檔案存至 static/uploads/ 資料夾中 並命名為 input.csv
# part 4
# 請求成功回傳 200 
def getCsvfromUrl(url):
    # part 1
    response = requests.get(url)

    # part 2
    if response.status_code != 200:
        return response.status_code

    # part 3
    with open('static/uploads/input.csv', 'wb') as file:
                for chunk in response:
                    file.write(chunk)
    # part 4
    return 200

# json into excel
# 將Json格式的資料轉換成excel
# part 1
# 將json格式的資料轉換成指定為檔名output.xlsx的檔案中，檔案位置會在指定專案跟目錄下static/xlsxFile/的資料夾中
# part 2
# 將輸出的excel中 若 '區'  欄位為 '需要手動處理' 就更改顏色為紅色
# part 3
# 在輸出成excel時，對輸出的中文檔名做格式轉換，若不轉換，當輸出檔名含有中文字元素時會出亂碼
# outputFileName 為將輸入的檔案名稱去除原有副檔名後更新為指定xlsx的副檔名，在放入其變數中
# part 4
# 回傳xlsx的檔案
def jsonIntoExcel(readerInJson,filename):
    
    # part 1
    # put json into excel
    pd.DataFrame(readerInJson).to_excel("static/xlsxFile/output.xlsx", index=False)  
            
    # part 2
    # update excel error address color
    path = "static/xlsxFile/output.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    # from the active attribute 
    sheet_obj = wb_obj.active
    max_row=sheet_obj.max_row

    for i in range(2,max_row+1):
        if sheet_obj.cell(row=i,column=7).value == '需要手動處理':
            sheet_obj.cell(row=i,column=7).font = Font(color = 'FFFF0000')
    wb_obj.save(path)
    
    # part 3
    # return xlsx file format setting
    outputFileName = quote(Path(filename).stem  + ".xlsx")
    # outputFileName = filename  + ".xlsx"
    xlsxFile = send_file("static/xlsxFile/output.xlsx", download_name= outputFileName , as_attachment=True)
    xlsxFile.headers['Content-Disposition']  = "; filename*=utf-8''{}".format(outputFileName)
    
    # part 4
    return xlsxFile


# reader processing into json format
################################
# index                      序號
# name                       姓名
# cellPhone              手機號碼
# telephone              坐機號碼
# province                     省
# city                         市
# area                         區
# addressDetail          詳細地址
# consignmentArray      託寄物名稱
# paymentMethod          付款方式
# orderCode                訂單號
# SKUArray                   SKU
# UPCArray                   UPC
# productCount              數量
# orderTotalPrice       訂單總金額
# orderTime              下單時間
# deliveryNote           配送備註
################################
# part 1
# 先初始化欲輸出欄位資訊為空陣列
#part 2
# 將所有資訊reader以一筆一筆的方式做讀取，之後將讀取到的資訊放入對應陣列中
#part 3
# 將放置在陣列中的資料轉換成json的格式
# part 4
# 回傳 json 格式的資料
def readerIntoJson(reader,consignment):
    #part 1
    # init empty array
    index = []
    name = []
    cellPhone = []
    telephone = []
    province = []
    city = []
    area = []
    addressDetail = []
    consignmentArray = []
    paymentMethod = []
    orderCode = []
    SKUArray =[]
    UPCArray = []
    productCount =[]
    orderTotalPrice = []
    orderTime = []
    deliveryNote = []

    #part 2
    count = 1
    for rows in reader:
        if len(rows) != 32:
            # input 格式錯誤 以每一筆資料長度是否為32判斷
             return 404
        
        # 分割地址 API
        oringeAddress = rows[13].replace(" ","") 
        respAddressSplit = addressSplitAPI(oringeAddress)

        # 外部分割地址API 發生錯誤
        if respAddressSplit == 501 :
            return 501

        # insert csv data into array
        index.append(count)
        name.append(rows[9])
        cellPhone.append(rows[25])
        telephone.append(rows[25])

        # input 資料有地址為空
        if respAddressSplit == 403 or respAddressSplit == 405:
            province.append("")
            city.append("")
            area.append("需要手動處理")
            addressDetail.append(oringeAddress)
        else:
            # 比對API回傳詳細地址與原地址是否一致
            checkAddress = oringeAddress.partition(respAddressSplit[3])
            if checkAddress[2] == respAddressSplit[0]:
                province.append(respAddressSplit[1])
                city.append(respAddressSplit[2])
                area.append(respAddressSplit[3])
                addressDetail.append(respAddressSplit[0])
            else:
                province.append("")
                city.append("")
                area.append("需要手動處理")
                addressDetail.append(oringeAddress)

        consignmentArray.append(consignment)
        paymentMethod.append(whitchPaymentmethod(rows[14]))
        orderCode.append(rows[6])
        SKUArray.append(rows[18])
        UPCArray.append(rows[18])
        productCount.append(rows[20])
        orderTotalPrice.append(rows[22])
        orderTime.append(time.strftime("%Y/%m/%d %H:%M:%S",time.localtime(time.time())))
        deliveryNote.append(rows[26])
        count += 1

    #part 3
    # put array data into json
    dataInJson = {
        "序號":index,
        "姓名":name,
        "手機號碼":cellPhone,
        "座機號碼":telephone,
        "省":province,
        "市":city,
        "區":area,
        "詳細地址":addressDetail,
        "托寄物名稱":consignmentArray,
        "付款方式":paymentMethod,
        "訂單號":orderCode,
        "SKU":SKUArray,
        "UPC":UPCArray,
        "數量":productCount,
        "訂單總金額":orderTotalPrice,
        "下單時間":orderTime,
        "配送備註":deliveryNote
    }

    #part 4
    return dataInJson

# file type condiction
# 判斷輸入檔案是否為規定的檔案格式 
# ALLOWED_EXTENSIONS 可以再新增可傳入的檔案格式
def allowedFileType(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS

# call splitAPI and Split response into array
# part 1
# 呼叫外部 API 做香港地址的分割 呼叫後放入變數response中
# part 2
# 將呼叫後的回傳值 response 更換成json格式放入 dataUnSplit，之後根據回傳值分別放入dataSplit(分割地址陣列) 和 dataDetail(詳細地址)
# 創建splitArray 第0位元素放 dataDetail(詳細地址)
# 用迴圈將dataSplit(分割地址陣列) 取出 真正的分割地址字串 放入 splitArray 的第1、2、3位元素
# part 3 
# 回傳 splitArray 陣列 ， 回傳陣列格式為 splitArray = ['詳細地址','省','市','區']
def addressSplitAPI(addr):
    # part 1
    response = requests.post(
        "https://htm.sf-express.com/sf-service-owf-web/service/order/orderAddressSplit?lang=tc&region=hk&translate=tc", 
        json={"address":"陳先生 "+ addr}
    )
    
    if response.status_code != 200:
        # 外部分割地址API 發生錯誤
        return 501

    # part 2
    dataUnSplit = response.json()
    dataSplit = dataUnSplit["result"]["originDestRegions"]
    dataDetail = dataUnSplit["result"]["site"]
    if dataDetail != "":
        splitArray = ["","","",""]
        splitArray[0] = dataDetail

        addressCount = 1
        for split in dataSplit:
            splitArray[addressCount] = split["name"]
            addressCount += 1
    else:
        # input 資料有地址為空
        return 403

    for splitElement in splitArray:
        if splitElement == "":
            # input 資料有地址為空，省、市、區
            return 405


    # part 3
    return splitArray

# choose witch Paymentmethod
# 付款方式的顯示字串判斷轉換
# 判斷input的paymethod付款方式為Y或N，Y為到付，N為寄付月結，若不為Y或N則輸出paymethod為空
def whitchPaymentmethod(paymethod):

    if paymethod == "Y":
        paymethod = "到付"
    elif paymethod == "N":
        paymethod = "寄付月結"
    else:
        paymethod = ""

    return paymethod